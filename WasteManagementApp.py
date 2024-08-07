import serial
import telebot
from tkinter import *
from time import strftime
from tkinter.ttk import *
from openpyxl import Workbook, load_workbook

class WasteManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Waste Management System')
        self.root.attributes('-fullscreen', True)
        self.root.configure(bg='white')

        self.ser = serial.Serial(port='/dev/ttyUSB0', baudrate=9600, parity=serial.PARITY_NONE,
                                 stopbits=serial.STOPBITS_ONE, bytesize=serial.EIGHTBITS, timeout=1)
        self.tb = telebot.TeleBot('YOUR_TOKEN_HERE')
        self.channel_id = 'YOUR_ID_HERE'

        self.setup_variables()
        self.setup_ui()
        self.start_clock()
        self.update_sensor_data()
        self.update_tong_bin()
        self.update_data()

    def setup_variables(self):
        self.arr_str = ["Num", "Date", "Time", "T/No", "G.w(kg)", "T.w(kg)", "N.w(kg)"]
        self.nama_kawasan = ["AREA1", "AREA2", "AREA3", "AREA4", "AREA5", "AREA6", "AREA7", "AREA8", "AREA9"]
        self.total_weights = [0.0] * len(self.nama_kawasan)
        self.current_weight = 0.0
        self.bin_count = 1
        self.tong_num = StringVar()
        self.selected_area = StringVar()
        self.row = 2

    def setup_ui(self):
        self.lbl = Label(self.root, font=('calibri', 100, 'bold'), background='white', foreground='black')
        self.lbl1 = Label(self.root, font=('calibri', 100, 'bold'), background='white', foreground='blue')
        self.lbl2 = Label(self.root, font=('calibri', 100, 'bold'), background='white', foreground='black')
        self.lbl3 = Label(self.root, font=('calibri', 100, 'bold'), background='white', foreground='black')
        self.tong = Label(self.root, font=('calibri', 100, 'bold'), background='white', foreground='black')
        self.total_lbl = Label(self.root, font=('calibri', 100, 'bold'), background='white', foreground='black')
        self.ayat = Label(self.root, font=('calibri', 100, 'bold'), background='white', foreground='black')
        self.dummy = Label(self.root, font=('calibri', 150, 'bold'), background='white', foreground='black')
        self.tong.config(text="Bin N0:")
        self.ayat.config(text="T.W: ")
        
        self.radiobuttons = []
        for idx, area in enumerate(self.nama_kawasan):
            rb = Radiobutton(self.root, variable=self.selected_area, value=area, text=area)
            self.radiobuttons.append(rb)

        self.tong_entry = Entry(self.root, textvariable=self.tong_num, font=('calibre', 80, 'bold'))
        self.save_button = Button(self.root, text="Save Data", command=self.submit_data, width=35)
        self.report_button = Button(self.root, text="Send Report", command=self.send_report, width=30)

        self.place_widgets()

    def place_widgets(self):
        self.lbl.place(x=50, y=50)
        self.lbl1.place(x=50, y=200)
        self.lbl2.place(x=50, y=350)
        self.lbl3.place(x=50, y=500)
        self.tong.place(x=50, y=650)
        self.tong_entry.place(x=500, y=630)
        for idx, rb in enumerate(self.radiobuttons):
            rb.place(x=100 + (idx % 2) * 150, y=850 + (idx // 2) * 30)
        self.total_lbl.place(x=700, y=850)
        self.ayat.place(x=400, y=850)
        self.save_button.place(x=500, y=800)
        self.report_button.place(x=200, y=800)

    def start_clock(self):
        self.update_time()
        self.root.after(1000, self.start_clock)

    def update_time(self):
        string = strftime("%I:%M:%S %p %d:%m:%y")
        self.lbl.config(text=string)

    def update_sensor_data(self):
        if self.ser.inWaiting() > 0:
            x = self.ser.readline().decode("utf-8").split()
            self.current_weight = float(x[1])
            self.lbl1.config(text=f"{self.current_weight} KG")
            net_weight = self.current_weight - 13.4
            self.lbl3.config(text=f"Waste Weight: {net_weight:.2f} KG")
        self.root.after(1000, self.update_sensor_data)

    def update_tong_bin(self):
        self.lbl2.config(text="Bin Weight: 13.4KG")
        self.root.after(1000, self.update_tong_bin)

    def update_data(self):
        self.root.after(1000, self.update_data)

    def submit_data(self):
        area = self.selected_area.get()
        if not area:
            return

        wb_path = f"/home/sailcott/Weight/{area}.xlsx"
        wb = load_workbook(wb_path)
        sheet = wb.active

        date = strftime("%d/%m/%Y")
        time = strftime("%I:%M:%S %p")
        total_weight = self.current_weight - 13.4

        data = [self.row - 1, date, time, self.tong_num.get(), self.current_weight, 13.4, total_weight]
        for col, value in enumerate(data, 1):
            sheet.cell(self.row, col).value = value

        self.total_weights[self.nama_kawasan.index(area)] += total_weight
        self.total_lbl.config(text=f"{self.total_weights[self.nama_kawasan.index(area)]:.2f}")

        self.tong_num.set("")
        self.row += 1
        wb.save(wb_path)

    def send_report(self):
        area = self.selected_area.get()
        if not area:
            return

        wb_path = f"/home/sailcott/Weight/{area}.xlsx"
        wb = load_workbook(wb_path)
        sheet = wb.active
        sheet.cell(self.row, 1).value = "Total Net Weight"
        sheet.cell(self.row, 7).value = self.total_weights[self.nama_kawasan.index(area)]
        wb.save(wb_path)

        self.total_weights[self.nama_kawasan.index(area)] = 0
        self.total_lbl.config(text="0.00")
        self.tb.send_message(self.channel_id, text=f"Hello guys!!\nNow is {strftime('%d/%m/%Y %I:%M %p')}\nand this is today's report")
        self.tb.send_document(self.channel_id, document=open(wb_path, 'rb'))
        self.row = 2
        urm = wb['Sheet']
        for row in urm.iter_rows():
            for cell in row:
                cell.value = None
        wb.save(wb_path)
        self.init_excel_file(wb_path)

    def init_excel_file(self, wb_path):
        wb = load_workbook(wb_path)
        sheet = wb.active
        for col, title in enumerate(self.arr_str, 1):
            sheet.cell(1, col).value = title
        wb.save(wb_path)

if __name__ == "__main__":
    root = Tk()
    app = WasteManagementApp(root)
    root.bind("<Escape>", exit)
    root.mainloop()
